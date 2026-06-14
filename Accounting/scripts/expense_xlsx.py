"""経費明細台帳の Excel(.xlsx)生成(scripts 層・openpyxl + Pillow)。

ledger/draft の内容(サイドカー相当)+ クラウド経費の明細番号 + 証憑サムネイルを一覧化する。
将来はクラウド経費から取り込んだ過去分も同じ行形式(相関キー/明細番号で突合)で追記できる。
core は zero-dep のため Excel/画像処理はここ(scripts/)に置く。
"""

from __future__ import annotations

import io
from pathlib import Path

# (列ヘッダ, row dict のキー)。先頭「証憑」は画像専用列(key=None)。
COLUMNS: list[tuple[str, str | None]] = [
    ("証憑", None),
    ("明細番号", "mf_number"),
    ("日付", "date"),
    ("支払先", "payee"),
    ("経費科目", "ex_item"),
    ("税区分", "excise"),
    ("金額", "amount"),
    ("通貨", "currency"),
    ("レート", "fx_rate"),
    ("円換算", "jpy_amount"),
    ("登録番号", "invoice_number"),
    ("内容", "description"),
    ("相関キー", "correlation_key"),
    ("ファイル名", "filename"),
    ("状態", "mf_status"),
    ("MF-ID", "mf_transaction_id"),
]
_WIDTHS = [22, 10, 12, 24, 26, 12, 11, 7, 7, 11, 14, 42, 18, 30, 10, 24]


def _thumbnail(path: str | Path, max_w: int = 140):
    """証憑画像を幅 max_w に縮小した PNG(BytesIO)と表示サイズを返す。"""
    from PIL import Image

    with Image.open(path) as im:
        im = im.convert("RGB")
        w, h = im.size
        h2 = max(1, round(h * max_w / w))
        im = im.resize((max_w, h2))
        buf = io.BytesIO()
        im.save(buf, format="PNG")
    buf.seek(0)
    return buf, (max_w, h2)


def build_xlsx(rows, out_path, *, images=None, row_key="correlation_key", sheet_title="経費明細"):
    """明細行(dict のリスト)から xlsx を生成する。`images` は {row[row_key]: 画像パス}。"""
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    images = images or {}
    out_path = Path(out_path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_fill = PatternFill("solid", fgColor="4472C4")
    for c, (title, _key) in enumerate(COLUMNS, start=1):
        cell = ws.cell(1, c, title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = _WIDTHS[c - 1]
    ws.freeze_panes = "A2"

    for i, row in enumerate(rows, start=2):
        for c, (_title, key) in enumerate(COLUMNS, start=1):
            if key is None:
                continue
            ws.cell(i, c, row.get(key) or "")
            ws.cell(i, c).alignment = Alignment(vertical="top", wrap_text=(key == "description"))
        img_path = images.get(row.get(row_key))
        if not (img_path and Path(img_path).is_file()):
            continue
        try:
            buf, (w, h) = _thumbnail(img_path)
        except (OSError, ValueError):
            continue
        xi = XLImage(buf)
        xi.width, xi.height = w, h
        ws.add_image(xi, f"A{i}")
        ws.row_dimensions[i].height = max(h * 0.75, 60)

    wb.save(out_path)
    return out_path
