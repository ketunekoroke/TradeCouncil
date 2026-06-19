"""scripts/expense_pipeline.py + ac expense CLI(з„ЎгғҚгғғгғҲгғҜгғјгӮҜгғ»SharePoint/PIL жіЁе…Ҙ)гҖӮ"""

import datetime
import json
import os

import pytest

from core import config, ingest
from core.moneyforward import ProductConfig
from scripts import expense_pipeline as ep


@pytest.fixture(autouse=True)
def _expense_env(monkeypatch, tmp_path):
    """дҪңжҘӯй ҳеҹҹгӮ’ tmp гҒ«еӣәе®ҡгҖӮе®ҹ .env / Teams URL гӮ’йҒ®ж–ӯгҒ—гҖҒйҖҡзҹҘгҒҜ no-op гҒ«гҒҷгӮӢгҖӮ"""
    monkeypatch.setenv("EXPENSE_VAR_DIR", str(tmp_path / "expense"))
    monkeypatch.setattr(config, "_root_env", lambda: {})
    monkeypatch.setattr(config, "_settings_local_env", lambda: {})
    for key in list(os.environ):
        if key.startswith(("MONEYFORWARD", "TEAMS")):
            monkeypatch.delenv(key, raising=False)
    yield


def _write_raw(name: str, data: bytes = b"%PDF-1.4 fake"):
    raw = ep.sub("raw")
    raw.mkdir(parents=True, exist_ok=True)
    (raw / name).write_bytes(data)


def _write_sidecar(raw_name: str, **fields):
    base = {
        "source_file": raw_name, "date": "2026-05-30", "payee": "JRжқұж—Ҙжң¬",
        "amount": "1000", "currency": "JPY", "description": "зү№жҖҘеҲё",
    }
    base.update(fields)
    ep.sub("extracted").mkdir(parents=True, exist_ok=True)
    ep.sidecar_path(raw_name).write_text(json.dumps(base, ensure_ascii=False), encoding="utf-8")


def _seed_usage():
    from core import refdata

    idx = refdata.aggregate_usage(
        [{"ex_item_name": "ж—…иІ»дәӨйҖҡиІ»", "excise_name": "иӘІзЁҺд»•е…Ҙ10%", "partner_name": "JRжқұж—Ҙжң¬"}]
    )
    ep.save_usage(idx)


def _write_split_spec(raw_name, parts, mode="explicit"):
    ep.sub("split").mkdir(parents=True, exist_ok=True)
    spec = {"source_file": raw_name, "mode": mode, "parts": parts}
    ep.split_spec_path(raw_name).write_text(json.dumps(spec, ensure_ascii=False), encoding="utf-8")


def _fake_split(src, jobs):
    """еҗ„гӮёгғ§гғ–гҒ®еҮәеҠӣе…ҲгҒ«жңҖе°Ҹ PDF гӮ’жӣёгҒҸ(pypdf гӮ’дҪҝгӮҸгҒӘгҒ„жіЁе…Ҙз”Ё)гҖӮ"""
    from pathlib import Path

    out = []
    for p, _pages in jobs:
        p = Path(p)
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_bytes(b"%PDF-1.4 " + p.name.encode())
        out.append(p)
    return out


def _pc_expense():
    return ProductConfig(
        product="expense", label="гӮҜгғ©гӮҰгғүзөҢиІ»", enabled=True, client_id="c", client_secret="s",
        authorize_url=None, token_url="https://t", redirect_uri=None, scopes=["transaction:write"],
        api_base="https://e/api", offices_url="https://e/api/external/v1/offices",
    )


def _seed_usage_with_ids():
    from core import refdata

    idx = refdata.aggregate_usage([
        {"ex_item": {"id": "EI_TAXI", "name": "ж—…иІ»дәӨйҖҡиІ»"},
         "dr_excise": {"id": "EX10", "long_name": "иӘІзЁҺд»•е…Ҙ 10%"}, "remark": "JRжқұж—Ҙжң¬"},
    ])
    ep.save_usage(idx)


def test_register_drafts_dry_run_then_confirm():
    _seed_usage_with_ids()
    _write_raw("a.pdf")
    _write_sidecar("a.pdf", ex_item="ж—…иІ»дәӨйҖҡиІ»", excise="иӘІзЁҺд»•е…Ҙ 10%")
    ep.process_all()
    pc = _pc_expense()

    res = ep.register_drafts(pc, confirm=False)  # гғүгғ©гӮӨгғ©гғі: йҖҒдҝЎгҒ—гҒӘгҒ„
    assert res["dry_run"] is True and not res["registered"]
    assert res["skipped"][0]["preview"]["ex_item_id"] == "EI_TAXI"
    assert res["skipped"][0]["preview"]["иЁјжҶ‘ж·»д»ҳ"] is True

    seen = {}

    def fake_create(pc, office_id, body, access_token=None):
        seen.update(body=body, office=office_id)
        return {"ex_transaction": {"id": "EXT999"}}

    res2 = ep.register_drafts(
        pc, confirm=True, create_fn=fake_create,
        get_office_id_fn=lambda pc, access_token: "of1", access_token="tok",
    )
    assert res2["registered"][0]["mf_id"] == "EXT999"
    assert seen["body"]["ex_transaction"]["ex_item_id"] == "EI_TAXI"
    assert "receipt_input" in seen["body"]["ex_transaction"]  # иЁјжҶ‘ж·»д»ҳ(йӣ»еёіжі•)
    led = ingest.Ledger.load(ep.ledger_path())
    assert led.entries[0].mf_status == "registered"
    assert led.entries[0].mf_transaction_id == "EXT999"

    res3 = ep.register_drafts(  # еҶҚе®ҹиЎҢ: зҷ»йҢІжёҲгҒҝгҒҜеҜҫиұЎеӨ–
        pc, confirm=True, create_fn=fake_create,
        get_office_id_fn=lambda pc, access_token: "of1", access_token="tok",
    )
    assert not res3["registered"]


def test_register_foreign_includes_memo_and_payee_remark():
    from core import refdata

    idx = refdata.aggregate_usage([
        {"ex_item": {"id": "EI_M", "name": "дјҡиӯ°иІ»"},
         "dr_excise": {"id": "EX_X", "long_name": "еҜҫиұЎеӨ–"}, "remark": "AWS"},
    ])
    ep.save_usage(idx)
    _write_raw("b.pdf")
    _write_sidecar("b.pdf", payee="AWS", amount="32.10", currency="USD",
                   description="EC2", ex_item="дјҡиӯ°иІ»", excise="еҜҫиұЎеӨ–", fx_rate="150")
    ep.process_all()
    seen = {}

    def fake_create(pc, office_id, body, access_token=None):
        seen["body"] = body
        return {"id": "X"}

    ep.register_drafts(
        _pc_expense(), confirm=True, create_fn=fake_create,
        get_office_id_fn=lambda pc, access_token: "of1", access_token="tok",
    )
    tx = seen["body"]["ex_transaction"]
    assert tx["remark"].startswith("AWS")  # еә—еҗҚгӮ’е…Ҳй ӯ
    assert "memo" in tx and "150" in tx["memo"]  # гғ¬гғјгғҲйҒ©з”ЁгғЎгғў
    assert tx["value"] == 32.1 and tx["currency"] == "USD"


def test_register_notifies_operations():
    _seed_usage_with_ids()
    _write_raw("a.pdf")
    _write_sidecar("a.pdf", ex_item="ж—…иІ»дәӨйҖҡиІ»", excise="иӘІзЁҺд»•е…Ҙ 10%")
    ep.process_all()
    calls = []

    def fake_notify(channel, title, message, *, facts=None, severity="info"):
        calls.append({"channel": channel, "facts": facts, "severity": severity})
        return True

    def fake_create(pc, oid, body, access_token=None):
        return {"ex_transaction": {"id": "X", "number": 950}}

    ep.register_drafts(
        _pc_expense(), confirm=True, create_fn=fake_create,
        get_office_id_fn=lambda pc, access_token: "of1", access_token="tok", notify_fn=fake_notify,
    )
    assert calls and calls[0]["channel"] == "operations"
    assert calls[0]["facts"]["жҳҺзҙ°з•ӘеҸ·"] == "950"
    assert calls[0]["facts"]["ж”Ҝжү•е…Ҳ"] == "JRжқұж—Ҙжң¬"
    assert calls[0]["facts"]["иЁјжҶ‘"].startswith("ж·»д»ҳгҒӮгӮҠ")


def test_register_skips_when_ex_item_id_unresolved():
    _seed_usage()  # еҗҚеүҚгҒ®гҒҝ(ID з„ЎгҒ—)
    _write_raw("a.pdf")
    _write_sidecar("a.pdf")
    ep.process_all()
    res = ep.register_drafts(_pc_expense(), confirm=False)
    assert res["skipped"] and "иІ»зӣ®IDжңӘи§Јжұә" in res["skipped"][0]["reason"]


def test_clean_inbox_only_registered():
    _seed_usage_with_ids()
    _write_raw("a.pdf")
    _write_sidecar("a.pdf", ex_item="ж—…иІ»дәӨйҖҡиІ»", excise="иӘІзЁҺд»•е…Ҙ 10%")
    ep.process_all()
    assert not ep.clean_inbox(confirm=False)["skipped"]  # жңӘзҷ»йҢІ вҶ’ еҜҫиұЎгҒӘгҒ—
    ep.register_drafts(
        _pc_expense(), confirm=True, create_fn=lambda *a, **k: {"id": "X"},
        get_office_id_fn=lambda pc, access_token: "of1", access_token="tok",
    )
    res = ep.clean_inbox(confirm=False)  # зҷ»йҢІеҫҢ вҶ’ dry-run гҒ§еҜҫиұЎгҒ«еҮәгӮӢ
    assert res["dry_run"] and res["skipped"][0]["source_file"] == "a.pdf"
    deleted = []
    res2 = ep.clean_inbox(confirm=True, delete_fn=lambda name: deleted.append(name))
    assert deleted == ["a.pdf"] and res2["deleted"] == ["a.pdf"]


def test_prior_fy_range():
    assert ep.prior_fy_range(datetime.date(2026, 6, 14)) == ("2024-07-01", "2025-06-30")
    assert ep.prior_fy_range(datetime.date(2026, 7, 1)) == ("2025-07-01", "2026-06-30")


def test_default_refdata_range_spans_prior_fy_to_today():
    # еүҚжңҹй–Ӣе§ӢгҖңд»Ҡж—Ҙ(еүҚжңҹгҒҢз©әгҒ§гӮӮеҪ“жңҹгҒ®е®ҹзёҫгӮ’еӯҰзҝ’гҒ§гҒҚгӮӢ)гҖӮ
    assert ep.default_refdata_range(datetime.date(2026, 6, 14)) == ("2024-07-01", "2026-06-14")


def test_expense_remote_paths():
    # drive зӣҙдёӢзӣёеҜҫ(root/workspace гӮ’д»ҳгҒ‘гҒӘгҒ„)гҖӮconfig гҒҢз©әгҒӘгӮүж—ўе®ҡгҖӮеүҚеҫҢгҒ® / гҒҜйҷӨеҺ»гҖӮ
    cfg = {"expense": {"inbox": "Expense/Inbox"}}
    assert ep.expense_remote(cfg, "inbox", "X") == "Expense/Inbox"
    assert ep.expense_remote({}, "inbox", "Expense/Inbox") == "Expense/Inbox"
    assert ep.expense_remote({"expense": {"master": "/Foo/Bar/"}}, "master", "X") == "Foo/Bar"


def test_run_refdata_aggregates_and_saves():
    pc = ProductConfig(
        product="expense", label="гӮҜгғ©гӮҰгғүзөҢиІ»", enabled=True, client_id="c", client_secret="s",
        authorize_url=None, token_url="https://t", redirect_uri=None, scopes=["transaction:write"],
        api_base="https://e/api", offices_url="https://e/api/external/v1/offices",
    )

    def fake_list(pc, *, date_from, date_to, access_token):
        assert (date_from, date_to) == ("2024-07-01", "2025-06-30")
        return [{"ex_item_name": "йҖҡдҝЎиІ»", "excise_name": "иӘІзЁҺд»•е…Ҙ10%", "partner_name": "AWS"}]

    summary = ep.run_refdata(pc, date_from="2024-07-01", date_to="2025-06-30", list_fn=fake_list)
    assert summary["n"] == 1
    assert ep.usage_path().is_file()
    assert ep.load_usage().suggest("AWS").ex_item == "йҖҡдҝЎиІ»"


def test_pending_extract_when_no_sidecar():
    _write_raw("a.pdf")
    res = ep.process_all()
    assert res["pending_extract"] == ["a.pdf"]
    assert not res["processed"]


def test_process_creates_draft_and_ledger():
    _seed_usage()
    _write_raw("a.pdf")
    _write_sidecar("a.pdf")
    res = ep.process_all()
    assert len(res["processed"]) == 1
    assert (ep.sub("processed") / "2026-05-30_JRжқұж—Ҙжң¬.pdf").exists()
    assert (ep.sub("processed") / "2026-05-30_JRжқұж—Ҙжң¬_original.pdf").exists()
    drafts = ep.list_drafts()
    assert drafts[0]["ex_transaction"]["ex_item_name"] == "ж—…иІ»дәӨйҖҡиІ»"
    assert drafts[0]["correlation_key"] in drafts[0]["ex_transaction"]["remark"]
    led = ingest.Ledger.load(ep.ledger_path())
    assert len(led.entries) == 1 and led.entries[0].ex_item == "ж—…иІ»дәӨйҖҡиІ»"


def test_image_box_passed_to_processor():
    _seed_usage()
    _write_raw("b.jpg", b"\xff\xd8\xff fake")
    _write_sidecar("b.jpg", crop_box=[1, 2, 30, 40])
    seen = {}

    def fake_proc(src, primary, original, box):
        seen["box"] = box
        from pathlib import Path

        Path(primary).write_bytes(b"x")
        Path(original).write_bytes(b"y")
        return {"primary": str(primary), "original": str(original), "cropped": box is not None}

    ep.process_all(process_image=fake_proc)
    assert seen["box"] == (1, 2, 30, 40)


def test_exact_duplicate_skipped_on_rerun():
    _seed_usage()
    _write_raw("a.pdf")
    _write_sidecar("a.pdf")
    ep.process_all()
    res2 = ep.process_all()
    assert res2["skipped"] and res2["skipped"][0]["reason"].startswith("е®Ңе…ЁйҮҚиӨҮ")
    assert not res2["processed"]


def test_similar_duplicate_needs_approval_then_overwrites():
    _seed_usage()
    _write_raw("a.pdf", b"content-A")
    _write_sidecar("a.pdf", amount="1000")
    ep.process_all()
    # еҗҢгҒҳж—Ҙд»ҳ+ж”Ҝжү•е…Ҳ+йҮ‘йЎҚгҒ гҒҢдёӯиә«гҒҢйҒ•гҒҶ(ж’®гӮҠзӣҙгҒ—)вҶ’ similar йҮҚиӨҮгҖӮ
    _write_raw("a.pdf", b"content-B-different-bytes")
    res = ep.process_all(approve_overwrite=False)  # approver гҒӘгҒ— вҶ’ дёҠжӣёгҒҚгҒӣгҒҡ skip
    assert res["skipped"] and "жңӘжүҝиӘҚ" in res["skipped"][0]["reason"]
    res2 = ep.process_all(approve_overwrite=True)
    assert res2["processed"] and res2["processed"][0]["overwrote"] is True
    led = ingest.Ledger.load(ep.ledger_path())
    assert len(led.entries) == 1
    assert led.entries[0].superseded  # дёҠжӣёгҒҚеұҘжӯҙгҒҢж®ӢгӮӢ


def test_check_compliance_scan_drafts_flags_errors():
    from scripts import check_compliance

    # usage гӮ’ж’’гҒӢгҒӘгҒ„ вҶ’ иІ»зӣ®жңӘзўәе®ҡ вҶ’ гӮІгғјгғҲ error вҶ’ scan_drafts гҒҢжӢҫгҒҶгҖӮ
    _write_raw("x.pdf")
    _write_sidecar("x.pdf", payee="еҗҚгӮӮз„ЎгҒҚеә—", description="дёҚжҳҺгҒӘе“Ғзӣ®")
    ep.process_all()
    problems = check_compliance.scan_drafts()
    assert problems and any("жңӘзўәе®ҡ" in p for p in problems)


def test_export_xlsx_from_ledger():
    import pytest

    pytest.importorskip("openpyxl")
    pimage = pytest.importorskip("PIL.Image")
    _seed_usage_with_ids()
    raw = ep.sub("raw")
    raw.mkdir(parents=True, exist_ok=True)
    pimage.new("RGB", (120, 160), (200, 200, 200)).save(raw / "r.png")
    _write_sidecar("r.png", payee="JRжқұж—Ҙжң¬", ex_item="ж—…иІ»дәӨйҖҡиІ»", excise="иӘІзЁҺд»•е…Ҙ 10%")
    ep.process_all()
    out = ep.export_xlsx()
    assert out.is_file()
    import openpyxl

    ws = openpyxl.load_workbook(out).active
    assert ws.cell(1, 4).value == "ж”Ҝжү•е…Ҳ"
    assert ws.cell(2, 4).value == "JRжқұж—Ҙжң¬"
    assert len(ws._images) == 1  # иЁјжҶ‘гӮөгғ гғҚгӮӨгғ«


def test_split_pdfs_dry_run_then_confirm():
    _write_raw("multi.pdf", b"%PDF-1.4 fake")
    _write_split_spec("multi.pdf", [{"pages": [1], "suffix": "a"}, {"pages": [2], "suffix": "b"}])

    dry = ep.split_pdfs(confirm=False, page_count_fn=lambda p: 2, split_fn=_fake_split)
    assert dry["dry_run"] and dry["skipped"][0]["parts"] == ["multi_a.pdf", "multi_b.pdf"]
    assert not (ep.sub("raw") / "multi_a.pdf").exists()  # гғүгғ©гӮӨгғ©гғігҒ§гҒҜжӣёгҒӢгҒӘгҒ„

    res = ep.split_pdfs(confirm=True, page_count_fn=lambda p: 2, split_fn=_fake_split)
    assert res["split"][0]["parts"] == ["multi_a.pdf", "multi_b.pdf"]
    assert (ep.sub("raw") / "multi_a.pdf").exists() and (ep.sub("raw") / "multi_b.pdf").exists()
    assert not (ep.sub("raw") / "multi.pdf").exists()  # еҺҹжң¬гҒҜ raw гҒӢгӮүйҖҖйҒҝ
    assert (ep.sub("split_src") / "multi.pdf").exists()  # йҖҖйҒҝе…ҲгҒ«ж®ӢгӮӢ(еҫ©е…ғеҸҜ)

    # еҶҚе®ҹиЎҢ: еҺҹжң¬гҒҜйҖҖйҒҝжёҲгҒҝ вҶ’ еҜҫиұЎгҒӘгҒ—(еҶӘзӯү)гҖӮ
    res2 = ep.split_pdfs(confirm=True, page_count_fn=lambda p: 2, split_fn=_fake_split)
    assert not res2["split"] and not res2["errors"]


def test_split_pdfs_catches_off_by_one():
    _write_raw("multi.pdf", b"%PDF-1.4 fake")
    _write_split_spec("multi.pdf", [{"pages": [3], "suffix": "a"}])  # е…Ё2гғҡгғјгӮёгҒ«3 вҶ’ зҜ„еӣІеӨ–
    res = ep.split_pdfs(confirm=True, page_count_fn=lambda p: 2, split_fn=_fake_split)
    assert res["errors"] and "зҜ„еӣІеӨ–" in res["errors"][0]["error"]
    assert (ep.sub("raw") / "multi.pdf").exists()  # еӨұж•—жҷӮгҒҜйҖҖйҒҝгҒ—гҒӘгҒ„


def test_split_pdfs_skips_existing_outputs():
    _write_raw("multi.pdf", b"%PDF-1.4 fake")
    _write_raw("multi_a.pdf", b"already")  # ж—ўгҒ«еҲҶеүІжёҲгҒҝзӣёеҪ“
    _write_split_spec("multi.pdf", [{"pages": [1], "suffix": "a"}])
    res = ep.split_pdfs(confirm=True, page_count_fn=lambda p: 1, split_fn=_fake_split)
    assert res["skipped"] and "ж—ўгҒ«еӯҳеңЁ" in res["skipped"][0]["reason"]


def test_split_then_process_each_part():
    # еҲҶеүІ вҶ’ еҗ„гғ‘гғјгғҲгҒ«жҠҪеҮәгӮөгӮӨгғүгӮ«гғј вҶ’ process гҒҢ1гғ¬гӮ·гғјгғҲгҒҡгҒӨеҮҰзҗҶгҖӮ
    _seed_usage()
    _write_raw("multi.pdf", b"%PDF-1.4 fake")
    _write_split_spec("multi.pdf", [{"pages": [1], "suffix": "a"}, {"pages": [2], "suffix": "b"}])
    ep.split_pdfs(confirm=True, page_count_fn=lambda p: 2, split_fn=_fake_split)
    _write_sidecar("multi_a.pdf", payee="SAN KYU", amount="500")
    _write_sidecar("multi_b.pdf", payee="2nd STREET", amount="800")
    res = ep.process_all()
    assert len(res["processed"]) == 2
    names = {p["processed_file"] for p in res["processed"]}
    assert any("SAN KYU" in n for n in names) and any("2nd STREET" in n for n in names)


def test_cli_split_dispatch(capsys, monkeypatch):
    from scripts import cli, pdfproc

    _write_raw("multi.pdf", b"%PDF-1.4 fake")
    _write_split_spec("multi.pdf", [{"pages": [1], "suffix": "a"}, {"pages": [2], "suffix": "b"}])
    monkeypatch.setattr(pdfproc, "page_count", lambda p: 2)  # pypdf гӮ’е‘јгҒ°гҒӣгҒӘгҒ„(dry-run)
    assert cli.main(["expense", "split"]) == 0
    out = capsys.readouterr().out
    assert "PDFеҲҶеүІ" in out and "DRY-RUN" in out and "multi_a.pdf" in out


def _tx_full(tx_id, *, date="2025-09-10", payee="Cafe del sol", ex_item="ж—…иІ»дәӨйҖҡиІ»",
             value=1500, has_file=True):
    tx = {
        "id": tx_id, "number": int(tx_id), "recognized_at": date,
        "ex_item": {"id": "EI_OLD", "name": ex_item},
        "dr_excise": {"id": "EX10", "long_name": "иӘІзЁҺд»•е…Ҙ10%"},
        "remark": payee, "value": value, "currency": "JPY",
    }
    if has_file:
        tx["mf_file"] = {"id": f"f{tx_id}", "name": "r.jpg",
                         "content_type": "image/jpeg", "byte_size": 3}
    return tx


def _office(pc, access_token):
    return "of1"


def test_import_past_downloads_and_seeds_ledger():
    txs = [
        _tx_full("100", date="2025-09-10"),                  # еңЁFY + иЁјжҶ‘
        _tx_full("200", date="2025-10-01", has_file=False),  # зҙҷ(иЁјжҶ‘гҒӘгҒ—)
        _tx_full("300", date="2024-08-01"),                  # жңҹеӨ–(еүҚжңҹ)
    ]

    def fake_list(pc, office_id, *, date_from, date_to, access_token, max_pages):
        return [t for t in txs if date_from <= t["recognized_at"] <= date_to]

    dl = []

    def fake_dl(pc, office_id, tx_id, *, access_token):
        dl.append(tx_id)
        return (b"img", "image/jpeg")

    res = ep.import_past(
        _pc_expense(), date_from="2025-07-01", date_to="2026-06-15",
        list_fn=fake_list, download_fn=fake_dl, get_office_id_fn=_office, access_token="tok",
    )
    assert {r["tx_id"] for r in res["imported"]} == {"100", "200"}  # 300 гҒҜжңҹеӨ–
    assert res["no_file"] == ["200"] and res["downloaded"] == 1 and dl == ["100"]
    assert (ep.past_dir() / "past_100.jpg").exists()
    assert ep.past_snapshot_path("100").is_file()
    led = ingest.Ledger.load(ep.ledger_path())
    e = next(x for x in led.entries if x.receipt_id == "past_100")
    assert e.mf_status == "imported" and e.mf_transaction_id == "100" and e.mf_number == "100"

    # еҶҚе®ҹиЎҢ: DLжёҲгҒҝ(гӮөгӮӨгӮәдёҖиҮҙ)гҒ§гӮ№гӮӯгғғгғ—=еҶӘзӯүгҖӮ
    dl.clear()
    res2 = ep.import_past(
        _pc_expense(), date_from="2025-07-01", date_to="2026-06-15",
        list_fn=fake_list, download_fn=fake_dl, get_office_id_fn=_office, access_token="tok",
    )
    assert dl == [] and any("DLжёҲгҒҝ" in s["reason"] for s in res2["skipped"])


def test_revise_past_dry_run_then_confirm():
    _seed_usage_with_ids()  # ж—…иІ»дәӨйҖҡиІ»вҶ’EI_TAXI, иӘІзЁҺд»•е…Ҙ 10%вҶ’EX10
    tx = _tx_full("100", ex_item="йҖҡдҝЎиІ»")  # MF гҒ® OCR гҒҢиӘӨгҒЈгҒҰйҖҡдҝЎиІ»

    def fake_list(pc, office_id, *, date_from, date_to, access_token, max_pages):
        return [tx]

    ep.import_past(
        _pc_expense(), date_from="2025-07-01", date_to="2026-06-15", list_fn=fake_list,
        download_fn=lambda *a, **k: (b"img", "image/jpeg"),
        get_office_id_fn=_office, access_token="tok",
    )
    # Claude еҶҚиӘӯиҫј: жӯЈгҒ—гҒ„иІ»зӣ®/зЁҺеҢәеҲҶгҒ®гӮөгӮӨгғүгӮ«гғј(past еҸ—й ҳгғ•гӮЎгӮӨгғ«еҗҚгҒ§)гҖӮ
    _write_sidecar("past_100.jpg", payee="JRжқұж—Ҙжң¬", amount="1500",
                   ex_item="ж—…иІ»дәӨйҖҡиІ»", excise="иӘІзЁҺд»•е…Ҙ 10%", date="2025-09-10")

    res = ep.revise_past(_pc_expense(), confirm=False)  # гғүгғ©гӮӨгғ©гғі
    assert res["dry_run"] and not res["revised"]
    prev = next(s for s in res["skipped"] if s.get("changes"))
    assert any(c["field"] == "ex_item" and c["after"] == "ж—…иІ»дәӨйҖҡиІ»" for c in prev["changes"])

    seen = {}

    def fake_update(pc, office_id, tx_id, body, access_token=None):
        seen.update(tx_id=tx_id, body=body)
        return {"id": tx_id}

    res2 = ep.revise_past(
        _pc_expense(), confirm=True, update_fn=fake_update,
        get_office_id_fn=_office, access_token="tok",
    )
    assert res2["revised"] and res2["revised"][0]["tx_id"] == "100"
    assert seen["body"]["ex_transaction"]["ex_item_id"] == "EI_TAXI"
    assert "receipt_input" not in seen["body"]["ex_transaction"]  # еҶҚгӮўгғғгғ—гғӯгғјгғүгҒ—гҒӘгҒ„
    led = ingest.Ledger.load(ep.ledger_path())
    e = next(x for x in led.entries if x.receipt_id == "past_100")
    assert e.mf_status == "revised" and e.revised_at and e.ex_item == "ж—…иІ»дәӨйҖҡиІ»"

    # еҶҚе®ҹиЎҢ: гӮ№гғҠгғғгғ—гӮ·гғ§гғғгғҲжӣҙж–°жёҲгҒҝ вҶ’ е·®еҲҶгӮјгғӯ(дәҢйҮҚPUTйҳІжӯў)гҖӮ
    res3 = ep.revise_past(
        _pc_expense(), confirm=True, update_fn=fake_update,
        get_office_id_fn=_office, access_token="tok",
    )
    assert res3["no_change"] == ["100"] and not res3["revised"]


def test_revise_past_skips_no_file_entry():
    _seed_usage_with_ids()
    tx = _tx_full("400", ex_item="йҖҡдҝЎиІ»", has_file=False)  # зҙҷ(иЁјжҶ‘гҒӘгҒ—)
    ep.import_past(
        _pc_expense(), date_from="2025-07-01", date_to="2026-06-15",
        list_fn=lambda *a, **k: [tx], download_fn=lambda *a, **k: (b"", ""),
        get_office_id_fn=_office, access_token="tok",
    )
    res = ep.revise_past(_pc_expense(), confirm=False)
    assert any("иЁјжҶ‘гҒӘгҒ—" in s["reason"] for s in res["skipped"]) and not res["revised"]


def test_export_xlsx_embeds_past_receipt():
    import io

    pytest.importorskip("openpyxl")
    pimage = pytest.importorskip("PIL.Image")
    buf = io.BytesIO()
    pimage.new("RGB", (60, 80), (180, 180, 180)).save(buf, format="PNG")
    png = buf.getvalue()
    tx = _tx_full("100")
    tx["mf_file"].update(name="r.png", byte_size=len(png))

    ep.import_past(
        _pc_expense(), date_from="2025-07-01", date_to="2026-06-15",
        list_fn=lambda *a, **k: [tx], download_fn=lambda *a, **k: (png, "image/png"),
        get_office_id_fn=_office, access_token="tok",
    )
    assert (ep.past_dir() / "past_100.png").exists()
    out = ep.export_xlsx()
    import openpyxl

    ws = openpyxl.load_workbook(out).active
    payees = [ws.cell(r, 4).value for r in range(2, ws.max_row + 1)]
    assert "Cafe del sol" in payees  # йҒҺеҺ»еҲҶгҒҢеҸ°еёігҒ«иЎҢгҒЁгҒ—гҒҰијүгӮӢ
    assert len(ws._images) >= 1  # past/ гҒӢгӮүиЁјжҶ‘гӮөгғ гғҚгӮӨгғ«гӮ’еҹӢиҫј


def test_export_xlsx_dedup_and_foreign_jpy():
    pytest.importorskip("openpyxl")
    tx = _tx_full("100", value=1000)
    tx["currency"], tx["jpyrate"] = "THB", 5.0  # еӨ–иІЁ(THB)гғ»гғ¬гғјгғҲ5.0
    ep.import_past(
        _pc_expense(), date_from="2025-07-01", date_to="2026-06-15",
        list_fn=lambda *a, **k: [tx], download_fn=lambda *a, **k: (b"img", "image/jpeg"),
        get_office_id_fn=_office, access_token="tok",
    )
    # еҗҢдёҖ MF еҸ–еј•(id=100)гҒ®еҸӨгҒ„ registered йҮҚиӨҮгӮ’еҸ°еёігҒ«ж··е…ҘгҒ•гҒӣгӮӢгҖӮ
    led = ingest.Ledger.load(ep.ledger_path())
    led.upsert(ingest.LedgerEntry(
        receipt_id="2026-03-14_old", content_hash="h", date="2026-03-14", payee="OLD",
        amount="1000", currency="THB", mf_status="registered", mf_transaction_id="100",
        correlation_key="AC-old",
    ))
    led.save(ep.ledger_path())

    out = ep.export_xlsx(embed_images=False)
    import openpyxl

    ws = openpyxl.load_workbook(out).active
    mfids = [ws.cell(r, 16).value for r in range(2, ws.max_row + 1)]  # MF-ID еҲ—
    assert mfids.count("100") == 1  # йҮҚиӨҮжҺ’йҷӨ: past_ еҒҙ1иЎҢгҒ®гҒҝ
    row = next(r for r in range(2, ws.max_row + 1) if ws.cell(r, 16).value == "100")
    assert str(ws.cell(row, 10).value) == "5000"  # еҶҶжҸӣз®—=valueГ—rate=1000Г—5.0
    assert ws.cell(row, 4).value == "Cafe del sol"  # past_ еҒҙ(OLD гҒ§гҒҜгҒӘгҒ„)


def test_cli_import_past_and_revise_past_dispatch(capsys, monkeypatch):
    from scripts import cli
    from scripts import expense_pipeline as ep_mod

    monkeypatch.setattr(ep_mod.oauth, "get_access_token", lambda pc: "tok")
    monkeypatch.setattr(ep_mod.mf_expense_api, "get_office_id", _office)
    monkeypatch.setattr(
        ep_mod.mf_expense_api, "list_my_ex_transactions",
        lambda pc, office_id, **k: [_tx_full("100")],
    )
    monkeypatch.setattr(
        ep_mod.mf_expense_api, "download_ex_transaction_receipt",
        lambda pc, office_id, tx_id, **k: (b"img", "image/jpeg"),
    )
    # гӮҜгғ©гӮҰгғүзөҢиІ»гҒ®иЁӯе®ҡгӮ’ ready гҒ«иҰӢгҒӣгӮӢ(import-past/revise-past гҒҜ load_product гӮ’йҖҡгӮӢ)гҖӮ
    monkeypatch.setattr(cli, "_err", lambda exc: str(exc))
    from core import moneyforward as mf
    monkeypatch.setattr(mf, "load_product", lambda product: _pc_expense())

    assert cli.main(["expense", "import-past"]) == 0
    out = capsys.readouterr().out
    assert "йҒҺеҺ»еҲҶеҸ–иҫј" in out
    assert cli.main(["expense", "revise-past"]) == 0  # гӮөгӮӨгғүгӮ«гғјз„ЎгҒ—вҶ’policy-only / dry-run
    out = capsys.readouterr().out
    assert "йҒҺеҺ»еҲҶиЈңжӯЈ" in out and "DRY-RUN" in out


def test_cli_status_and_process_dispatch(capsys):
    from scripts import cli

    _seed_usage()
    _write_raw("a.pdf")
    _write_sidecar("a.pdf")
    assert cli.main(["expense", "process", "--yes"]) == 0
    out = capsys.readouterr().out
    assert "еҮҰзҗҶ 1" in out
    assert cli.main(["expense", "status"]) == 0
    out = capsys.readouterr().out
    assert "еҸ°еёі 1 д»¶" in out


def test_fetch_masters_merges_ids_into_usage():
    from core import refdata

    # еӯҰзҝ’жёҲгҒҝ usage(йҖҡдҝЎиІ»=LEARNED)гӮ’зЁ®гҒ«гҒҷгӮӢгҖӮ
    idx = refdata.aggregate_usage(
        [{"ex_item": {"name": "йҖҡдҝЎиІ»", "id": "LEARNED"}, "remark": "AWS"}]
    )
    ep.save_usage(idx)

    ex_items = [{"id": "E1", "name": "йҖҡдҝЎиІ»"}, {"id": "E2", "name": "ж”Ҝжү•жүӢж•°ж–ҷ"}]
    excises = [{"id": "T8", "name": "иӘІзЁҺд»•е…Ҙ 8%"}]
    summary = ep.fetch_masters(
        _pc_expense(),
        access_token="tok",
        ex_items_fn=lambda pc, oid, *, access_token=None: ex_items,
        excises_fn=lambda pc, oid, *, access_token=None: excises,
        get_office_id_fn=lambda pc, *, access_token=None: "of1",
    )
    assert summary["ex_items"] == 2 and summary["excises"] == 1
    assert summary["added_ex_item"] == 1  # йҖҡдҝЎиІ»гҒҜеӯҰзҝ’жёҲгҒҝвҶ’ж”Ҝжү•жүӢж•°ж–ҷгҒ®гҒҝиҝҪеҠ 
    assert summary["added_excise"] == 1
    u = ep.load_usage()
    assert u.ex_item_id("йҖҡдҝЎиІ»") == "LEARNED"  # еӯҰзҝ’жёҲгҒҝгҒҜдҝқжҢҒ
    assert u.ex_item_id("ж”Ҝжү•жүӢж•°ж–ҷ") == "E2"  # жңӘеӯҰзҝ’иІ»зӣ®гӮ’и§ЈжұәеҸҜиғҪгҒ«
    assert u.excise_id("иӘІзЁҺд»•е…Ҙ 8%") == "T8"  # 8% гӮ’и§ЈжұәеҸҜиғҪгҒ«


def test_fetch_masters_handles_missing_office():
    summary = ep.fetch_masters(
        _pc_expense(),
        access_token="tok",
        ex_items_fn=lambda *a, **k: [],
        excises_fn=lambda *a, **k: [],
        get_office_id_fn=lambda pc, *, access_token=None: None,
    )
    assert summary.get("error")


def test_is_var_core_filters_state_from_images():
    # зҠ¶ж…ӢгҒ®ж ё(еҶҚзҸҫдёҚеҸҜ)= TrueгҖҒиЁјжҶ‘з”»еғҸ/дёҖжҷӮзү© = FalseгҖӮ
    assert ep._is_var_core("ledger.json")
    assert ep._is_var_core("extracted/2026-05-30_x.json")
    assert ep._is_var_core("refdata/expense_usage.json")
    assert ep._is_var_core("drafts/r.json")
    assert ep._is_var_core("murc_2026.xls")
    assert ep._is_var_core("past/past_abc.mf.json")  # MFзҸҫеҖӨгӮ№гғҠгғғгғ—гӮ·гғ§гғғгғҲ=зҠ¶ж…Ӣ
    assert not ep._is_var_core("processed/2026-05-30_x.pdf")
    assert not ep._is_var_core("past/past_abc.jpg")  # з”»еғҸжң¬дҪ“
    assert not ep._is_var_core("raw/scan.pdf")


def test_plan_var_sync_core_only_excludes_images():
    local = {"ledger.json": 100.0, "processed/a.jpg": 100.0, "raw/x.pdf": 100.0}
    remote = {"drafts/d.json": 200.0}
    full = dict(ep.plan_var_sync(local, remote, core_only=False))
    assert full["ledger.json"] == "push"
    assert full["processed/a.jpg"] == "push"  # йҒ йҡ”гҒ«з„ЎгҒ„ вҶ’ push
    assert full["drafts/d.json"] == "pull"  # гғӯгғјгӮ«гғ«гҒ«з„ЎгҒ„ вҶ’ pull
    core = dict(ep.plan_var_sync(local, remote, core_only=True))
    assert core == {"ledger.json": "push", "drafts/d.json": "pull"}  # з”»еғҸ/raw гҒҜеҜҫиұЎеӨ–


def test_sync_var_orchestration_injected():
    pushes, pulls = [], []
    res = ep.sync_var(
        connect=lambda base: ("drv", "Expense/Var/expense", "tok"),
        local_index_fn=lambda b: {"ledger.json": 100.0, "processed/a.jpg": 100.0},
        remote_index_fn=lambda d, rb, t: {
            "ledger.json": (50.0, {"id": "r1"}),  # local newer вҶ’ push
            "drafts/d.json": (200.0, {"id": "r2"}),  # local з„ЎгҒ— вҶ’ pull
        },
        push_fn=lambda d, lp, rf, t: pushes.append(rf),
        pull_fn=lambda d, it, rf, lp, t: pulls.append((rf, it)),
    )
    assert set(res["pushed"]) == {"ledger.json", "processed/a.jpg"}
    assert res["pulled"] == ["drafts/d.json"]
    assert res["remote"] == "Expense/Var/expense"
    assert "Expense/Var/expense/ledger.json" in pushes
    assert pulls == [("Expense/Var/expense/drafts/d.json", {"id": "r2"})]  # remote item гӮ’ pull гҒ«жёЎгҒҷ


def test_sync_var_refreshes_token_on_401():
    calls = {"push": 0, "refresh": 0}

    def push(d, lp, rf, t):
        calls["push"] += 1
        if t == "old":  # ж—§гғҲгғјгӮҜгғігҒҜеӨұеҠ№ вҶ’ 401
            raise RuntimeError("Microsoft Graph HTTP 401: token is expired")

    def refresh():
        calls["refresh"] += 1
        return "new"

    res = ep.sync_var(
        connect=lambda base: ("drv", "Expense/Var/expense", "old"),
        local_index_fn=lambda b: {"ledger.json": 100.0},
        remote_index_fn=lambda d, rb, t: {},
        push_fn=push,
        pull_fn=lambda *a, **k: None,
        token_refresh=refresh,
    )
    assert res["pushed"] == ["ledger.json"]
    assert calls["refresh"] == 1  # 401 жӨңзҹҘгҒ§еҶҚеҸ–еҫ—
    assert calls["push"] == 2  # еӨұж•—1 + еҶҚи©ҰиЎҢжҲҗеҠҹ1
    assert res["token_refreshes"] == 1


def test_sync_var_time_based_token_refresh():
    # гӮҜгғӯгғғгӮҜгӮ’йҖІгӮҒгҖҒд»¶ж•°гҒҢе°‘гҒӘгҒҸгҒҰгӮӮжҷӮй–“зөҢйҒҺгҒ§е…ҲеӣһгӮҠгғҲгғјгӮҜгғіжӣҙж–°гҒҢиө·гҒҚгӮӢгҒ“гҒЁгӮ’жӨңиЁјгҖӮ
    ticks = iter([0.0, 0.0, 5000.0])  # 2д»¶зӣ®гҒ®еүҚгҒ«40еҲҶи¶…(2400s)зөҢйҒҺ
    refreshed = {"n": 0}

    def clock():
        try:
            return next(ticks)
        except StopIteration:
            return 9999.0

    def refresh():
        refreshed["n"] += 1
        return "new"

    res = ep.sync_var(
        connect=lambda base: ("drv", "Expense/Var/expense", "old"),
        local_index_fn=lambda b: {"a.txt": 1.0, "b.txt": 1.0},
        remote_index_fn=lambda d, rb, t: {},
        push_fn=lambda *a, **k: None,
        pull_fn=lambda *a, **k: None,
        token_refresh=refresh,
        refresh_interval_sec=2400.0,
        refresh_every=0,  # д»¶ж•°гғҷгғјгӮ№гҒҜз„ЎеҠ№еҢ–гҒ—жҷӮй–“гғҷгғјгӮ№гҒ®гҒҝжӨңиЁј
        clock=clock,
    )
    assert res["token_refreshes"] >= 1
    assert refreshed["n"] >= 1
