"""scripts/expense_xlsx.py: 明細台帳 xlsx 生成(openpyxl/Pillow 未導入なら skip)。"""

import pytest

from scripts import expense_xlsx


def test_build_xlsx_with_thumbnail(tmp_path):
    pytest.importorskip("openpyxl")
    pimage = pytest.importorskip("PIL.Image")
    img = tmp_path / "r.png"
    pimage.new("RGB", (200, 300), (10, 20, 30)).save(img)
    rows = [{
        "mf_number": "946", "date": "2026-06-02", "payee": "sairee cottage restaurant",
        "ex_item": "接待飲食費(1人あたり10000円以下)", "excise": "対象外", "amount": "3100.00",
        "currency": "THB", "fx_rate": "4.90", "jpy_amount": "15190", "invoice_number": "",
        "description": "飲食(接待)", "correlation_key": "AC-202606-x", "filename": "r.png",
        "mf_status": "registered", "mf_transaction_id": "7gQV",
    }]
    out = tmp_path / "out.xlsx"
    expense_xlsx.build_xlsx(rows, out, images={"AC-202606-x": str(img)})
    assert out.is_file()

    import openpyxl

    ws = openpyxl.load_workbook(out).active
    assert ws.cell(1, 1).value == "証憑"
    assert ws.cell(1, 2).value == "明細番号"
    assert ws.cell(2, 2).value == "946"
    assert ws.cell(2, 4).value == "sairee cottage restaurant"
    assert len(ws._images) == 1  # 証憑サムネイルが埋め込まれている
